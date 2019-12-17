from openpyxl import load_workbook
from openpyxl import Workbook
import xlrd

wb = Workbook()


def open_edge_file():
    filename1 = "" # N1_List
    Data = xlrd.open_workbook(filename1).sheet_by_index(0)
    Data_col = Data.ncols
    Data_row = Data.nrows

    N1_List = []
    N1_List.append(Data)
    N1_List.append(Data_row)
    N1_List.append(Data_col)

    filename2 = "" # N2_List
    Data = xlrd.open_workbook(filename2).sheet_by_index(0)
    Data_col = Data.ncols
    Data_row = Data.nrows

    N2_List = []
    N2_List.append(Data)
    N2_List.append(Data_row)
    N2_List.append(Data_col)

    return N1_List, N2_List


def compare_edge(N1_List, N2_List):
    file_name = 'Edge_compare_result.xlsx'
    sheet1 = wb.active
    sheet2 = wb.active
    sheet3 = wb.active
    sheet1.title = 'Overlap edge list'
    sheet2.title = 'Edges in only Network 1'
    sheet2.title = 'Edges in only Network 2'


    Overlap_List = set(N1_List) & set(N2_List)
    Overlap_List = list(Overlap_List)
    print(Overlap_List)

    for row_index in range(len(Overlap_List)):
        sheet1.cell(row=row_index + 1, column=1).value = Overlap_List[row_index][0]
        sheet1.cell(row=row_index + 1, column=2).value = Overlap_List[row_index][1]

    # Edges in only Network 1
    Only_N1 = set(N1_List) - set(Overlap_List)
    Only_N1 = list(Only_N1)

    print(Only_N1)
    for row_index in range(len(Only_N1)):
        sheet2.cell(row=row_index + 1, column=1).value = Only_N1[row_index][0]
        sheet2.cell(row=row_index + 1, column=2).value = Only_N1[row_index][1]

    #==============================================================================================
    # Edges in only Network 2
    Only_N2 = set(N2_List) - set(Overlap_List)
    Only_N2 = list(Only_N2)
    print(Only_N2)

    for row_index in range(len(Only_N2)):
        sheet3.cell(row=row_index + 1, column=1).value = Only_N2[row_index][0]
        sheet3.cell(row=row_index + 1, column=2).value = Only_N2[row_index][1]
    wb.save(filename=file_name)

if __name__ == "__main__":
    compare_edge(open_edge_file())